import openpyxl
from app.models import Data
from app.utils import fake_data, get_fields_name
from app.loader import load


def parse(filepath, timestamper=fake_data):
    wookbook = openpyxl.load_workbook(filepath)
    worksheet = wookbook.active
    fields = get_fields_name(Data)[1:]  # cut primary key

    bulk_data = []
    for i in range(3, worksheet.max_row):
        row = [col[i].value for col in worksheet.iter_cols(2, worksheet.max_column)]
        #  cut two last empty cells and add some date
        dict_row = {
            k: v for k,v in zip(
                fields, row[:-2] + [timestamper()]
            )
        }
        bulk_data.append(dict_row)
    load(bulk_data)

if __name__ == "__main__":
    parse()
